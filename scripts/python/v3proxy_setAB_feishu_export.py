"""
v3proxy_setAB_feishu_export.py
Purpose: Build focused Set A vs Set B coefficient comparison outputs for
         D, D x SR, DrySM, and DrySM x SR; save local csv/xlsx/figures; and
         sync figure/value tables to Feishu Bitable.
"""

from __future__ import annotations

import json
import math
import os
import subprocess
from pathlib import Path
from typing import Any, Dict, Iterable, List

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.lines import Line2D
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOTDIR = Path(r"C:\YangSu\00_Project\CA_mechanism\regression_SR")
TABDIR = ROOTDIR / "output" / "tables"
FIGDIR = ROOTDIR / "output" / "figures"
XLSDIR = ROOTDIR / "output" / "spreadsheet"

BASE_TOKEN = "XsH1beeEMaMf0Os04IYcSz54nhh"
FIGURE_TABLE_NAME = "v3proxy_setAB_figures"
VALUES_TABLE_NAME = "v3proxy_setAB_values"

LARK_CLI = [
    "node",
    r"C:\Users\Lenovo\AppData\Roaming\npm\node_modules\@larksuite\cli\scripts\run.js",
]

ENV = os.environ.copy()
ENV["MSYS_NO_PATHCONV"] = "1"
ENV["LARK_CLI_NO_PROXY"] = "1"
for proxy_key in ("HTTPS_PROXY", "HTTP_PROXY", "https_proxy", "http_proxy"):
    ENV.pop(proxy_key, None)


CUT_ORDER = {"full": 1, "ii": 2, "iv": 3}
RAW_ORDER = {"full": 1, "v3pm10": 2, "hepm10": 3, "v3pre30": 4, "v3he": 5, "hema": 6}
SOURCE_ORDER = {"Set A baseline": 0, "GLEAM": 1, "SWSM": 2, "ERA5-Land": 3}
LAYER_ORDER = {"Baseline": 0, "Surface": 1, "Rootzone": 2}

CUT_LABELS = {
    "full": "Full",
    "ii": "V3pm10 + HEpm10",
    "iv": "V3pre30 + V3HE + HEMA",
}

RAW_LABELS = {
    "full": "Full",
    "v3pm10": "V3pm10",
    "hepm10": "HEpm10",
    "v3pre30": "V3pre30",
    "v3he": "V3HE",
    "hema": "HEMA",
}

SOURCE_LABELS = {
    "baseline": "Set A baseline",
    "gleam": "GLEAM",
    "swsm": "SWSM",
    "era": "ERA5-Land",
}

LAYER_LABELS = {
    "baseline": "Baseline",
    "surface": "Surface",
    "rootzone": "Rootzone",
}

TERM_FAMILY = {
    "D": "Shock coefficient",
    "drysm": "Shock coefficient",
    "D_x_ca": "Buffering interaction",
    "drysm_x_ca": "Buffering interaction",
}

TERM_LABEL = {
    "D": "D",
    "D_x_ca": "D x SR",
    "drysm": "DrySM",
    "drysm_x_ca": "DrySM x SR",
}

FAMILY_KEY = {
    "Shock coefficient": "shock",
    "Buffering interaction": "buffer",
}

COLOR_MAP = {
    "Set A baseline": "#616161",
    "GLEAM": "#2E7D32",
    "SWSM": "#E65100",
    "ERA5-Land": "#1565C0",
}

MARKER_MAP = {
    "Baseline": "s",
    "Surface": "o",
    "Rootzone": "^",
}

LEGEND_NOTE = (
    "Legend: color = Set A baseline or SM source; shape = baseline, surface, "
    "or rootzone; each point is a coefficient with 95% CI."
)

SUBTITLE = (
    "Reduced controls, strict common sample, grid + year FE; Set A provides the "
    "baseline D or D x SR row, Set B provides DrySM or DrySM x SR rows by source and layer."
)


def decode_output(raw: bytes) -> str:
    for enc in ("utf-8", "gbk"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def lark_api(
    method: str,
    path: str,
    data: Dict[str, Any] | None = None,
    file_path: Path | None = None,
) -> Dict[str, Any]:
    cmd = list(LARK_CLI) + ["api", method, path, "--as", "user"]
    if data is not None:
        cmd += ["--data", json.dumps(data, ensure_ascii=True)]
    if file_path is not None:
        rel_path = os.path.relpath(file_path, ROOTDIR)
        cmd += ["--file", f"file={rel_path}"]

    result = subprocess.run(
        cmd,
        capture_output=True,
        shell=False,
        env=ENV,
        cwd=str(ROOTDIR),
        check=False,
    )

    stdout = decode_output(result.stdout).strip()
    stderr = decode_output(result.stderr).strip()
    payload_raw = stdout or stderr
    if not payload_raw:
        raise RuntimeError("Empty response from lark-cli")

    try:
        payload = json.loads(payload_raw)
    except json.JSONDecodeError as exc:
        raise RuntimeError(payload_raw[:800]) from exc

    if payload.get("ok") is False:
        raise RuntimeError(payload_raw[:800])
    if payload.get("code") not in (0, None):
        raise RuntimeError(payload_raw[:800])
    return payload


def list_tables() -> List[Dict[str, Any]]:
    payload = lark_api("GET", f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables?page_size=100")
    return payload["data"]["items"]


def create_table(table_name: str) -> str:
    payload = {"table": {"name": table_name}}
    resp = lark_api("POST", f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables", payload)
    return resp["data"]["table_id"]


def delete_table(table_id: str) -> None:
    lark_api("DELETE", f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables/{table_id}")


def create_field(table_id: str, field_name: str, field_type: int) -> None:
    payload = {"field_name": field_name, "type": field_type}
    lark_api(
        "POST",
        f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables/{table_id}/fields",
        payload,
    )


def ensure_clean_table(table_name: str, fields: List[Dict[str, Any]]) -> str:
    for item in list_tables():
        if item["name"] == table_name:
            delete_table(item["table_id"])
            break

    table_id = create_table(table_name)
    for field in fields:
        create_field(table_id, field["field_name"], field["type"])
    return table_id


def batch_create_records(table_id: str, records: List[Dict[str, Any]], batch_size: int = 10) -> None:
    for start in range(0, len(records), batch_size):
        batch = records[start : start + batch_size]
        lark_api(
            "POST",
            f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables/{table_id}/records/batch_create",
            {"records": batch},
        )


def upload_bitable_attachment(file_path: Path, parent_type: str) -> str:
    payload = {
        "file_name": file_path.name,
        "parent_type": parent_type,
        "parent_node": BASE_TOKEN,
        "size": file_path.stat().st_size,
    }
    resp = lark_api("POST", "/open-apis/drive/v1/medias/upload_all", payload, file_path=file_path)
    return resp["data"]["file_token"]


def as_number(value: Any) -> float | int | None:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    return None


def build_core_dataframe() -> pd.DataFrame:
    df = pd.read_csv(TABDIR / "v3proxy_results_long.csv")
    df = df[
        (df["module"] == "drought")
        & (df["ctrl_version"] == "reduced")
        & (df["model_set"].isin(["SetA", "SetB"]))
        & (df["term"].isin(["D", "D_x_ca", "drysm", "drysm_x_ca"]))
        & (df["coef_present"] == 1)
    ].copy()

    df["coef_family"] = df["term"].map(TERM_FAMILY)
    df["coefficient_label"] = df["term"].map(TERM_LABEL)
    df["cut_label"] = df["cut"].map(CUT_LABELS)
    df["raw_window_label"] = df["raw_window"].map(RAW_LABELS)
    df["series"] = df["source"].map(SOURCE_LABELS)
    df["shape_group"] = df["layer"].map(LAYER_LABELS)
    df["ci_lo"] = df["b"] - 1.96 * df["se"]
    df["ci_hi"] = df["b"] + 1.96 * df["se"]
    df["figure_key"] = df["coef_family"].map(FAMILY_KEY)
    df["figure_id"] = df["figure_key"] + "_" + df["cut"]
    df["sheet_name"] = df["figure_id"]
    df["figure_title"] = (
        df["coef_family"]
        + ": "
        + df["cut_label"]
        + " ("
        + df["coefficient_label"].where(df["model_set"] == "SetA", other=df["coefficient_label"])
        + ")"
    )
    df["display_label"] = df.apply(
        lambda row: (
            f"{row['raw_window_label']} | Set A baseline"
            if row["model_set"] == "SetA"
            else f"{row['raw_window_label']} | {row['series']} | {row['shape_group']}"
        ),
        axis=1,
    )
    df["sort_cut"] = df["cut"].map(CUT_ORDER)
    df["sort_raw"] = df["raw_window"].map(RAW_ORDER)
    df["sort_series"] = df["series"].map(SOURCE_ORDER)
    df["sort_layer"] = df["shape_group"].map(LAYER_ORDER)
    df["sort_family"] = df["coef_family"].map({"Shock coefficient": 1, "Buffering interaction": 2})

    ordered_cols = [
        "module",
        "model_set",
        "ctrl_version",
        "cut",
        "cut_label",
        "raw_window",
        "raw_window_label",
        "term",
        "coefficient_label",
        "coef_family",
        "figure_id",
        "sheet_name",
        "source",
        "series",
        "layer",
        "shape_group",
        "display_label",
        "b",
        "se",
        "p",
        "ci_lo",
        "ci_hi",
        "N",
        "r2",
    ]

    return df.sort_values(
        ["sort_family", "sort_cut", "sort_raw", "sort_series", "sort_layer"]
    )[ordered_cols].reset_index(drop=True)


def build_figure_index(core_df: pd.DataFrame) -> pd.DataFrame:
    fig_df = (
        core_df.groupby(["figure_id", "sheet_name", "cut", "cut_label", "coef_family"], as_index=False)
        .agg(row_count=("figure_id", "size"))
        .copy()
    )
    fig_df["figure_file"] = fig_df["figure_id"].apply(
        lambda x: str(FIGDIR / f"v3proxy_{x}.png")
    )
    fig_df["legend_note"] = LEGEND_NOTE
    fig_df["figure_title"] = fig_df.apply(
        lambda row: f"{row['coef_family']}: {row['cut_label']}", axis=1
    )
    return fig_df.sort_values(["coef_family", "cut"]).reset_index(drop=True)


def plot_figure(core_df: pd.DataFrame, cut: str, coef_family: str, out_path: Path) -> None:
    plot_df = core_df[(core_df["cut"] == cut) & (core_df["coef_family"] == coef_family)].copy()
    plot_df["sort_raw"] = plot_df["raw_window"].map(RAW_ORDER)
    plot_df["sort_series"] = plot_df["series"].map(SOURCE_ORDER)
    plot_df["sort_layer"] = plot_df["shape_group"].map(LAYER_ORDER)
    plot_df = plot_df.sort_values(["sort_raw", "sort_series", "sort_layer"]).reset_index(drop=True)
    plot_df["y"] = list(range(len(plot_df) - 1, -1, -1))

    xmin = plot_df["ci_lo"].min()
    xmax = plot_df["ci_hi"].max()
    xpad = max(0.01, (xmax - xmin) * 0.12)

    fig_height = max(4.8, 0.38 * len(plot_df) + 1.9)
    fig, ax = plt.subplots(figsize=(12.8, fig_height), dpi=300)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    ax.axvline(0, color="#6d6d6d", linestyle="--", linewidth=0.8, zorder=0)

    for _, row in plot_df.iterrows():
        color = COLOR_MAP[row["series"]]
        marker = MARKER_MAP[row["shape_group"]]
        ax.plot([row["ci_lo"], row["ci_hi"]], [row["y"], row["y"]], color=color, linewidth=1.5, zorder=2)
        ax.scatter(
            row["b"],
            row["y"],
            color=color,
            marker=marker,
            s=58,
            edgecolor="black",
            linewidth=0.35,
            zorder=3,
        )

    group_edges = (
        plot_df.groupby("raw_window_label", sort=False)["y"]
        .agg(["min", "max"])
        .reset_index(drop=True)
    )
    for _, edge in group_edges.iloc[:-1].iterrows():
        ax.axhline(edge["min"] - 0.5, color="#e0e0e0", linewidth=0.8, zorder=1)

    ax.set_yticks(plot_df["y"])
    ax.set_yticklabels(plot_df["display_label"], fontsize=9)
    ax.set_xlim(xmin - xpad, xmax + xpad)
    ax.set_xlabel("Coefficient (95% CI)", fontsize=10.5)
    ax.set_ylabel("")
    ax.set_title(f"{coef_family}: {CUT_LABELS[cut]}", fontsize=14, fontweight="bold", loc="left")
    ax.text(0, 1.01, SUBTITLE, transform=ax.transAxes, ha="left", va="bottom", fontsize=9.5, color="#444444")

    ax.grid(axis="x", color="#ededed", linewidth=0.8)
    ax.grid(axis="y", visible=False)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)

    color_handles = [
        Line2D([0], [0], marker="o", color="none", markerfacecolor=COLOR_MAP[label], markeredgecolor="black",
               markeredgewidth=0.3, markersize=7, label=label)
        for label in ["Set A baseline", "GLEAM", "SWSM", "ERA5-Land"]
    ]
    shape_handles = [
        Line2D([0], [0], marker=MARKER_MAP[label], color="black", markerfacecolor="white",
               markersize=7, linewidth=0, label=label)
        for label in ["Baseline", "Surface", "Rootzone"]
    ]

    leg1 = ax.legend(handles=color_handles, title="", loc="upper center", bbox_to_anchor=(0.33, -0.12),
                     ncol=4, frameon=False, fontsize=9)
    ax.add_artist(leg1)
    ax.legend(handles=shape_handles, title="", loc="upper center", bbox_to_anchor=(0.82, -0.12),
              ncol=3, frameon=False, fontsize=9)

    fig.text(0.01, 0.015, LEGEND_NOTE, fontsize=9, ha="left")
    plt.tight_layout(rect=(0, 0.05, 1, 0.97))
    fig.savefig(out_path, dpi=300, facecolor="white", bbox_inches="tight")
    plt.close(fig)


def save_plots(core_df: pd.DataFrame, fig_index: pd.DataFrame) -> None:
    FIGDIR.mkdir(parents=True, exist_ok=True)
    for _, row in fig_index.iterrows():
        plot_figure(
            core_df,
            cut=row["cut"],
            coef_family=row["coef_family"],
            out_path=Path(row["figure_file"]),
        )


def write_dataframe_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=sheet_name)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(row.tolist(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    numeric_cols = {"b", "se", "p", "ci_lo", "ci_hi", "r2"}
    integer_cols = {"N", "row_count"}
    for idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), *(len(str(v)) for v in df[col_name].fillna("").tolist()))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 28)
        if col_name in numeric_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).number_format = "0.0000"
        if col_name in integer_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).number_format = "0"


def add_plot_sheet(wb: Workbook, sheet_name: str, title_text: str, figure_path: Path) -> None:
    ws = wb.create_sheet(title=sheet_name)
    ws["A1"] = title_text
    ws["A2"] = SUBTITLE
    ws["A3"] = LEGEND_NOTE
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 28

    img = XLImage(str(figure_path))
    img.anchor = "A5"
    ws.add_image(img)


def build_workbook(core_df: pd.DataFrame, fig_index: pd.DataFrame, out_path: Path) -> None:
    XLSDIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    write_dataframe_sheet(wb, "coef_long", core_df)
    write_dataframe_sheet(wb, "figure_index", fig_index)

    for _, row in fig_index.iterrows():
        add_plot_sheet(
            wb=wb,
            sheet_name=row["sheet_name"],
            title_text=row["figure_title"],
            figure_path=Path(row["figure_file"]),
        )

    wb.save(out_path)


def build_figure_records(fig_index: pd.DataFrame, png_tokens: Dict[str, str]) -> List[Dict[str, Any]]:
    records = []
    for _, row in fig_index.iterrows():
        records.append(
            {
                "fields": {
                    "figure_id": row["figure_id"],
                    "cut": row["cut"],
                    "coef_family": row["coef_family"],
                    "figure_title": row["figure_title"],
                    "sheet_name": row["sheet_name"],
                    "row_count": int(row["row_count"]),
                    "legend_note": row["legend_note"],
                    "figure_png": [{"file_token": png_tokens[row["figure_id"]]}],
                }
            }
        )
    return records


def build_value_records(core_df: pd.DataFrame) -> List[Dict[str, Any]]:
    records = []
    for _, row in core_df.iterrows():
        records.append(
            {
                "fields": {
                    "figure_id": row["figure_id"],
                    "cut": row["cut"],
                    "raw_window": row["raw_window"],
                    "coef_family": row["coef_family"],
                    "model_set": row["model_set"],
                    "term": row["term"],
                    "coefficient_label": row["coefficient_label"],
                    "source": row["source"],
                    "layer": row["layer"],
                    "series": row["series"],
                    "shape_group": row["shape_group"],
                    "display_label": row["display_label"],
                    "b": as_number(row["b"]),
                    "se": as_number(row["se"]),
                    "p": as_number(row["p"]),
                    "ci_lo": as_number(row["ci_lo"]),
                    "ci_hi": as_number(row["ci_hi"]),
                    "N": int(row["N"]),
                    "r2": as_number(row["r2"]),
                }
            }
        )
    return records


def sync_feishu(fig_index: pd.DataFrame, core_df: pd.DataFrame) -> Dict[str, str]:
    figure_fields = [
        {"field_name": "figure_id", "type": 1},
        {"field_name": "cut", "type": 1},
        {"field_name": "coef_family", "type": 1},
        {"field_name": "figure_title", "type": 1},
        {"field_name": "sheet_name", "type": 1},
        {"field_name": "row_count", "type": 2},
        {"field_name": "legend_note", "type": 1},
        {"field_name": "figure_png", "type": 17},
    ]

    value_fields = [
        {"field_name": "figure_id", "type": 1},
        {"field_name": "cut", "type": 1},
        {"field_name": "raw_window", "type": 1},
        {"field_name": "coef_family", "type": 1},
        {"field_name": "model_set", "type": 1},
        {"field_name": "term", "type": 1},
        {"field_name": "coefficient_label", "type": 1},
        {"field_name": "source", "type": 1},
        {"field_name": "layer", "type": 1},
        {"field_name": "series", "type": 1},
        {"field_name": "shape_group", "type": 1},
        {"field_name": "display_label", "type": 1},
        {"field_name": "b", "type": 2},
        {"field_name": "se", "type": 2},
        {"field_name": "p", "type": 2},
        {"field_name": "ci_lo", "type": 2},
        {"field_name": "ci_hi", "type": 2},
        {"field_name": "N", "type": 2},
        {"field_name": "r2", "type": 2},
    ]

    figure_table_id = ensure_clean_table(FIGURE_TABLE_NAME, figure_fields)
    value_table_id = ensure_clean_table(VALUES_TABLE_NAME, value_fields)

    png_tokens = {
        row["figure_id"]: upload_bitable_attachment(Path(row["figure_file"]), "bitable_image")
        for _, row in fig_index.iterrows()
    }

    batch_create_records(figure_table_id, build_figure_records(fig_index, png_tokens))
    batch_create_records(value_table_id, build_value_records(core_df))

    return {
        "figure_table_id": figure_table_id,
        "value_table_id": value_table_id,
    }


def main() -> None:
    TABDIR.mkdir(parents=True, exist_ok=True)
    FIGDIR.mkdir(parents=True, exist_ok=True)
    XLSDIR.mkdir(parents=True, exist_ok=True)

    core_df = build_core_dataframe()
    fig_index = build_figure_index(core_df)

    core_csv = TABDIR / "v3proxy_setAB_core_coefficients.csv"
    fig_csv = TABDIR / "v3proxy_setAB_figure_index.csv"
    workbook_path = XLSDIR / "v3proxy_setAB_core_coefficients.xlsx"

    core_df.to_csv(core_csv, index=False, encoding="utf-8-sig")
    fig_index.to_csv(fig_csv, index=False, encoding="utf-8-sig")

    save_plots(core_df, fig_index)
    build_workbook(core_df, fig_index, workbook_path)

    sync_info = sync_feishu(fig_index, core_df)

    print(f"Saved coefficient csv: {core_csv}")
    print(f"Saved figure index csv: {fig_csv}")
    print(f"Saved workbook: {workbook_path}")
    print(f"Saved figures: {len(fig_index)}")
    print(
        "Feishu tables: "
        f"{FIGURE_TABLE_NAME}={sync_info['figure_table_id']}, "
        f"{VALUES_TABLE_NAME}={sync_info['value_table_id']}"
    )


if __name__ == "__main__":
    main()
